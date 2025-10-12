import pandas as pd
import openpyxl
from openpyxl import load_workbook
import os

# 输入和输出文件路径
input_file = '/Users/gndys/Documents/编程开发/#开发中ind/可行性 MVP/02/9.29 新/00-数据准备/任务分配表/任务分配表.xlsx'
output_file = '/Users/gndys/Documents/编程开发/#开发中ind/可行性 MVP/02/9.29 新/00-数据准备/任务分配表/任务分配表_整理后.xlsx'

# 读取工作簿并计算所有公式
print("正在读取并计算Excel公式...")
wb = load_workbook(input_file, data_only=True)

# 定义标准列名
standard_columns = [
    '负责人', '区域',
    '1月', '2月', '3月', 'Q1',
    '4月', '5月', '6月', 'Q2',
    '7月', '8月', '9月', 'Q3',
    '10月', '11月', '12月', 'Q4',
    '全年合计',
    '24年任务', '24年实际完成', '24年完成率', '任务增长率'
]

# 创建一个新的Excel writer
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    for sheet_name in wb.sheetnames:
        print(f"\n处理工作表: {sheet_name}")
        
        # 读取原始数据（使用data_only=True已经计算了公式）
        ws = wb[sheet_name]
        
        # 提取数据到列表
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))
        
        # 创建DataFrame，跳过前两行表头，使用标准列名
        df = pd.DataFrame(data[2:], columns=standard_columns)
        
        # 清理数据：移除完全为空的行
        df = df.dropna(how='all')
        
        # 将数值列转换为数值类型
        numeric_columns = [
            '1月', '2月', '3月', 'Q1',
            '4月', '5月', '6月', 'Q2',
            '7月', '8月', '9月', 'Q3',
            '10月', '11月', '12月', 'Q4',
            '全年合计',
            '24年任务', '24年实际完成', '24年完成率', '任务增长率'
        ]
        
        for col in numeric_columns:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')
        
        # 打印处理后的数据信息
        print(f"  - 数据行数: {len(df)}")
        print(f"  - 列数: {len(df.columns)}")
        print(f"  - 前5行:")
        print(df.head())
        
        # 写入到新的Excel文件
        df.to_excel(writer, sheet_name=sheet_name, index=False)

print(f"\n整理完成！输出文件: {output_file}")

# 再次读取并显示汇总信息
print("\n" + "="*60)
print("整理后的数据汇总:")
print("="*60)

for sheet_name in ['杭州', '金华', '湖州']:
    df = pd.read_excel(output_file, sheet_name=sheet_name)
    print(f"\n【{sheet_name}】")
    print(f"行数: {len(df)}, 列数: {len(df.columns)}")
    print(f"\n列名:")
    for i, col in enumerate(df.columns, 1):
        print(f"  {i}. {col}")
    
    print(f"\n数据预览:")
    print(df.head(3))
    print(f"\n负责人列表: {df['负责人'].dropna().unique().tolist()}")
    print(f"区域列表: {df['区域'].dropna().unique().tolist()}")

