import pandas as pd
import openpyxl

# 读取Excel文件
file_path = '/Users/gndys/Documents/编程开发/#开发中ind/可行性 MVP/02/9.29 新/00-数据准备/任务分配表/任务分配表.xlsx'

# 使用openpyxl查看原始结构
wb = openpyxl.load_workbook(file_path)
print("工作表列表:", wb.sheetnames)
print("\n" + "="*50)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n工作表: {sheet_name}")
    print(f"使用范围: {ws.dimensions}")
    print(f"\n前10行数据:")
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        print(f"第{i}行: {row}")
        if i >= 10:
            break
    print("\n" + "="*50)

# 使用pandas查看
print("\n使用pandas读取:")
for sheet_name in wb.sheetnames:
    print(f"\n工作表: {sheet_name}")
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    print(f"形状: {df.shape}")
    print(f"\n前15行:")
    print(df.head(15))

